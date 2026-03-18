"""
Generates sample_input.xlsx in the same directory as this script.

Usage:  python3 docs/samples/generate_sample_xlsx.py
Requires: openpyxl  (pip install openpyxl)
"""

from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

DATE_FMT = "DD/MM/YYYY HH:MM:SS"

# ---------------------------------------------------------------------------
# Column headers  (must match aliases in src/config.js)
# ---------------------------------------------------------------------------
HEADERS = [
    "Publication ID",
    "Episode ID",
    "Availability Mode",
    "Publication Date and Time",
    "Window Closure Date and Time",
    "Channel Label",
    "Channel ID",
    "Is Repeat?",
    "Is Primary?",
    "Sub-channel 1 Label",
    "Sub-channel 1 ID",
    "Sub-channel 2 Label",
]

# ---------------------------------------------------------------------------
# Sample rows  — None means leave the cell empty
# ---------------------------------------------------------------------------
ROWS = [
    # Broadcast, primary, two regional sub-channels
    {
        "Publication ID":            "PUB-20250315-BBC1-001",
        "Episode ID":                "b0abc1234",
        "Availability Mode":         "broadcast",
        "Publication Date and Time": datetime(2025, 3, 15, 20, 0, 0),
        "Window Closure Date and Time": None,
        "Channel Label":             "BBC One",
        "Channel ID":                "BBC One",
        "Is Repeat?":                False,
        "Is Primary?":               True,
        "Sub-channel 1 Label":       "BBC One Wales",
        "Sub-channel 1 ID":          "BBC One Wales",
        "Sub-channel 2 Label":       "BBC One Scotland",
    },
    # Broadcast repeat, no sub-channels
    {
        "Publication ID":            "PUB-20250315-ITV-001",
        "Episode ID":                "itv_ep_9876",
        "Availability Mode":         "broadcast",
        "Publication Date and Time": datetime(2025, 3, 15, 21, 0, 0),
        "Window Closure Date and Time": None,
        "Channel Label":             "ITV1",
        "Channel ID":                None,   # defaults to Channel Label
        "Is Repeat?":                True,
        "Is Primary?":               False,
        "Sub-channel 1 Label":       None,
        "Sub-channel 1 ID":          None,
        "Sub-channel 2 Label":       None,
    },
    # On-demand with window closure
    {
        "Publication ID":            "PUB-20250315-IPLAYER-001",
        "Episode ID":                "b0abc1234",
        "Availability Mode":         "onDemand",
        "Publication Date and Time": datetime(2025, 3, 15, 20, 0, 0),
        "Window Closure Date and Time": datetime(2025, 4, 15, 23, 59, 59),
        "Channel Label":             "BBC iPlayer",
        "Channel ID":                "iPlayer",
        "Is Repeat?":                True,
        "Is Primary?":               False,
        "Sub-channel 1 Label":       None,
        "Sub-channel 1 ID":          None,
        "Sub-channel 2 Label":       None,
    },
    # On-demand, no closure date (open-ended)
    {
        "Publication ID":            "PUB-20250316-C4OD-001",
        "Episode ID":                "c4_ep_5555",
        "Availability Mode":         "onDemand",
        "Publication Date and Time": datetime(2025, 3, 16, 9, 0, 0),
        "Window Closure Date and Time": None,
        "Channel Label":             "Channel 4",
        "Channel ID":                "C4",
        "Is Repeat?":                False,
        "Is Primary?":               True,
        "Sub-channel 1 Label":       None,
        "Sub-channel 1 ID":          None,
        "Sub-channel 2 Label":       None,
    },
    # Intentionally bad row — wrong Availability Mode value (tests validation)
    {
        "Publication ID":            "PUB-20250316-BAD-001",
        "Episode ID":                "bad_ep_0001",
        "Availability Mode":         "live",    # invalid — should be broadcast/onDemand
        "Publication Date and Time": datetime(2025, 3, 16, 18, 30, 0),
        "Window Closure Date and Time": None,
        "Channel Label":             "Sky One",
        "Channel ID":                None,
        "Is Repeat?":                "yes",
        "Is Primary?":               "no",
        "Sub-channel 1 Label":       None,
        "Sub-channel 1 ID":          None,
        "Sub-channel 2 Label":       None,
    },
]

# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Publications"

header_fill = PatternFill("solid", fgColor="2563EB")
header_font = Font(bold=True, color="FFFFFF", size=10)

for col_idx, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

for row_idx, row_data in enumerate(ROWS, start=2):
    for col_idx, header in enumerate(HEADERS, start=1):
        value = row_data.get(header)
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        if isinstance(value, datetime):
            cell.number_format = DATE_FMT

# Auto-size columns
for col in ws.columns:
    max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
    ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------
out_path = Path(__file__).parent / "sample_input.xlsx"
out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(out_path)
print(f"Saved: {out_path}")
